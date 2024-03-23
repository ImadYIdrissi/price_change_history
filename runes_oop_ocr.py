import cv2
import os
import numpy as np
import pytesseract

from io import BytesIO
from pathlib import Path


# from PIL import Image as PILImage
from typing import Tuple, List, Dict, Any
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# Define color constants for drawing on images
BLACK: Tuple[int, int, int] = (0, 0, 0)
WHITE: Tuple[int, int, int] = (255, 255, 255)
BLUE: Tuple[int, int, int] = (255, 0, 0)
GREEN: Tuple[int, int, int] = (0, 255, 0)
RED: Tuple[int, int, int] = (0, 0, 255)

ENLARGE_FACTOR: float = 1.5


# TODO : Customize training the 'Kamas' sign into 'K' or 'K€'
class TextElement:
    def __init__(
        self, contour: np.ndarray, offset: Tuple[int, int] = (0, 0)
    ) -> None:
        """
        Initialize a TextElement object with a contour and an optional offset.

        Args:
            contour: A numpy array containing contour points of the text
            element.
            offset: A tuple representing the offset to be applied to the
            contour's position, default is (0, 0).
        """
        self.x, self.y, self.w, self.h = cv2.boundingRect(array=contour)
        self.x += offset[0]
        self.y += offset[1]

    def _preprocess_image_1(
        self,
        img: np.ndarray,
    ) -> np.ndarray:
        """
        Preprocess the image for OCR by converting it to grayscale, applying
        Gaussian blur, adaptive thresholding, a morphological close operation,
        inverting the colors, and adding padding around the image to ensure
        text does not touch the borders.

        Args:
            img: The image to preprocess.

        Returns:
            The preprocessed and padded image.
        """
        # Enlarge the image
        enlarge_factor = ENLARGE_FACTOR
        padding = 10
        width = int(img.shape[1] * enlarge_factor)
        height = int(img.shape[0] * enlarge_factor)
        dim = (width, height)

        img = cv2.resize(img, dim, interpolation=cv2.INTER_AREA)

        # Convert to grayscale
        gray = cv2.cvtColor(src=img, code=cv2.COLOR_BGR2GRAY)

        # Apply Gaussian blur
        blurred = cv2.GaussianBlur(src=gray, ksize=(5, 5), sigmaX=0)

        # Use adaptive thresholding
        adaptive_thresh = cv2.adaptiveThreshold(
            src=blurred,
            maxValue=255,
            adaptiveMethod=cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            thresholdType=cv2.THRESH_BINARY,
            blockSize=11,
            C=2,
        )

        # Morphological operation
        kernel = cv2.getStructuringElement(shape=cv2.MORPH_RECT, ksize=(2, 2))
        morph = cv2.morphologyEx(
            src=adaptive_thresh, op=cv2.MORPH_CLOSE, kernel=kernel
        )

        # # Detect edges using Canny to find the contours of the black areas
        # edges = cv2.Canny(morph, 100, 200)

        # # Dilate the edges to extend them
        # kernel = np.ones((3, 3), np.uint8)
        # dilated_edges = cv2.dilate(edges, kernel, iterations=1)

        # # Superimpose the dilated edges onto the original image to extend
        # # the black regions
        # # Wherever the edges are white (255), we want to convert those pixels
        # # to black (0) in the original image
        # img[dilated_edges == 255] = 0
        # display_img(dilated_edges)
        # display_img(img)

        # display_img(inverted_morph)

        # TODO : Remove the double "lines" representing the contour
        # + the actual letter

        # Add padding
        padded_image = cv2.copyMakeBorder(
            morph,
            top=int(padding / 2),
            bottom=int(padding / 2),
            left=padding,
            right=padding,
            borderType=cv2.BORDER_CONSTANT,
            value=WHITE,
        )

        # Invert colors
        inverted = cv2.bitwise_not(padded_image)
        # display_img(inverted)

        # display_img(padded_image)
        return inverted

    # TODO : Add OCR function that uses PaddleOCR

    def ocr_text_with_pytesseract(
        self, original_img: np.ndarray
    ) -> Dict[str, Any]:
        """
        Perform OCR on a cropped and preprocessed part of the original image.
        Args:
            original_img: The original image to perform OCR on.
        Returns:
            A dictionary with keys 'original_cropped_img',
            'preprocessed_cropped_img' and 'text', containing the
            original image, the preprocessed image and the OCR-extracted text,
            respectively.
        """
        cropped_img = original_img[
            self.y : self.y + self.h, self.x : self.x + self.w
        ]
        # Preprocess the cropped image
        preprocessed_img = self._preprocess_image_1(img=cropped_img)
        # Apply pytesseract OCR on the preprocessed image
        text = pytesseract.image_to_string(
            image=preprocessed_img, config="--psm 7 --oem 3"
        ).strip()
        return {
            "original_cropped_img": cropped_img,
            "preprocessed_cropped_img": preprocessed_img,
            "text": text,
        }

    def draw(
        self,
        img: np.ndarray,
        display: bool = False,
        color: Tuple[int, int, int] = RED,
        window_name: str = "Text Element",
    ) -> np.ndarray:
        """
        Draw a bounding box around the text element on the image,
        optionally display it, and allow color customization.

        Args:
            img: The original image on which to draw.
            display: Flag indicating whether to display the image with the
            bounding box.
            color: The color of the bounding box (B, G, R).

        Returns:
            The image with the bounding box drawn around the text element.
        """
        # Draw a rectangle around the text element with the specified color
        cv2.rectangle(
            img=img,
            pt1=(self.x, self.y),
            pt2=(self.x + self.w, self.y + self.h),
            color=color,
            thickness=2,
        )

        if display:
            cv2.imshow(winname=window_name, mat=img)
            cv2.waitKey(
                delay=0
            )  # Wait for a key press to close the displayed window
            cv2.destroyAllWindows()

        return img


class Line(TextElement):
    def __init__(
        self,
        contour: np.ndarray,
        line_id: int,
        offset: Tuple[int, int] = (0, 0),
    ) -> None:
        """
        Initialize a Line object, inheriting from TextElement, to represent a
        line of text.

        Args:
            contour: A numpy array containing contour points of the line.
            line_id: An integer representing the unique identifier of the line.
            offset: A tuple representing the offset to be applied to the line's
                    position, default is (0, 0).
        """
        super().__init__(contour=contour, offset=offset)
        self.line_id = line_id
        self.phrases: List[Phrase] = []


class Phrase(TextElement):
    def __init__(
        self,
        contour: np.ndarray,
        line_ref: int,
        offset: Tuple[int, int] = (0, 0),
    ) -> None:
        """
        Initialize a Phrase object, inheriting from TextElement, to represent
        a phrase within a line.

        Args:
            contour: A numpy array containing contour points of the phrase.
            line_ref: An integer reference to the parent Line object.
            offset: A tuple representing the offset to be applied to the
                    phrase's position, default is (0, 0).
        """
        super().__init__(contour=contour, offset=offset)
        self.line_ref = line_ref
        self.words: List[Word] = []


class Word(TextElement):
    def __init__(
        self,
        contour: np.ndarray,
        phrase_ref: Phrase,
        offset: Tuple[int, int] = (0, 0),
    ) -> None:
        """
        Initialize a Word object, inheriting from TextElement, to represent a
        word within a phrase.

        Args:
            contour: A numpy array containing contour points of the word.
            phrase_ref: A reference to the parent Phrase object.
            offset: A tuple representing the offset to be applied to the word's
                    position, default is (0, 0).
        """
        super().__init__(contour=contour, offset=offset)
        self.phrase_ref = phrase_ref


def preprocess_image_for_contours(
    img_path: str,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Load an image from the given path and preprocess it to find contours.

    Args:
        img_path: The file path of the image.

    Returns:
        A tuple containing the original image and the thresholded image.
    """
    img = cv2.imread(filename=img_path)
    if img is None:
        raise ValueError("The image could not be loaded.")
    gray = cv2.cvtColor(src=img, code=cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(
        src=gray, thresh=150, maxval=255, type=cv2.THRESH_BINARY
    )
    return img, thresh


def find_contours(
    thresh: np.ndarray, kernel_size: Tuple[int, int], iterations: int
) -> List[np.ndarray]:
    """
    Find contours in the given thresholded image using a specified kernel size
    and number of iterations for dilation.

    Args:
        thresh: The thresholded image.
        kernel_size: A tuple specifying the size of the kernel used for
        dilation.
        iterations: The number of times dilation is applied.

    Returns:
        A list of contours found in the image.
    """
    kernel = cv2.getStructuringElement(shape=cv2.MORPH_RECT, ksize=kernel_size)
    dilate = cv2.dilate(src=thresh, kernel=kernel, iterations=iterations)
    contours, _ = cv2.findContours(
        image=dilate, mode=cv2.RETR_EXTERNAL, method=cv2.CHAIN_APPROX_SIMPLE
    )
    return contours


def detect_text_elements(img: np.ndarray, thresh: np.ndarray) -> List[Line]:
    """
    Detect and organize text elements (lines, phrases, words) in the given
    image based on the provided thresholded image.

    Args:
        img: The original image.
        thresh: The thresholded image.

    Returns:
        A list of Line objects, each containing their respective Phrase and
        Word objects.
    """

    def _get_vertical_position(contour: np.ndarray) -> int:
        _, y, _, _ = cv2.boundingRect(array=contour)
        return y

    def _get_horizontal_position(contour: np.ndarray) -> int:
        x, _, _, _ = cv2.boundingRect(array=contour)
        return x

    lines: List[Line] = []
    line_contours = find_contours(
        thresh=thresh, kernel_size=(70, 5), iterations=5
    )

    sorted_line_contours = sorted(line_contours, key=_get_vertical_position)

    for line_id, contour in enumerate(sorted_line_contours):
        line = Line(contour=contour, line_id=line_id)
        lines.append(line)

        line_thresh = thresh[
            line.y : line.y + line.h, line.x : line.x + line.w
        ]
        phrase_contours = find_contours(
            thresh=line_thresh, kernel_size=(20, 5), iterations=2
        )
        sorted_phrase_contours = sorted(
            phrase_contours, key=_get_horizontal_position
        )

        for phrase_contour in sorted_phrase_contours:
            phrase = Phrase(
                contour=phrase_contour,
                line_ref=line.line_id,
                offset=(line.x, line.y),
            )
            line.phrases.append(phrase)

            phrase_thresh = thresh[
                phrase.y : phrase.y + phrase.h, phrase.x : phrase.x + phrase.w
            ]
            word_contours = find_contours(
                thresh=phrase_thresh, kernel_size=(4, 5), iterations=2
            )
            sorted_word_contours = sorted(
                word_contours, key=_get_horizontal_position
            )

            for word_contour in sorted_word_contours:
                word = Word(
                    contour=word_contour,
                    phrase_ref=phrase,
                    offset=(phrase.x, phrase.y),
                )
                phrase.words.append(word)

    return lines


def display_img(mat: np.ndarray, tag: str = "") -> None:
    """
    Display an image with a given tag in the window title.

    Args:
        mat: The image matrix to display.
        tag: An optional tag to append to the window title.
    """
    cv2.imshow(winname="Image" + tag, mat=mat)
    cv2.waitKey(0)
    cv2.destroyAllWindows()


def add_image_to_excel_from_memory(
    image: np.ndarray, sheet, row: int, col: str
):
    is_success, buffer = cv2.imencode(".png", image)
    if is_success:
        # Directly use the buffer for creating an OpenpyxlImage
        image_stream = BytesIO(buffer)
        openpyxl_image = OpenpyxlImage(image_stream)
        cell = f"{col}{row}"
        sheet.add_image(openpyxl_image, cell)


# After all the images and text have been inserted into the worksheet
def adjust_column_widths(worksheet, min_width=10):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * ENLARGE_FACTOR / 2
        worksheet.column_dimensions[get_column_letter(column)].width = max(
            min_width, adjusted_width
        )


def add_results_to_excel(
    file_path: str,
    sheet_name: str,
    ocred_tuples: List[Tuple[np.ndarray, np.ndarray, str]],
    delete_file_if_exists: bool = False,
):
    file_path = Path(file_path)

    # Delete the file if it exists and deletion is requested
    if file_path.exists() and delete_file_if_exists:
        os.remove(file_path)

    # Load or create the workbook
    if not file_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # The workbook is new, so headers will go to the first row
        header_row = 1
    else:
        wb = load_workbook(filename=file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Check if the sheet is empty and headers should go to the
            # first row
            if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
                header_row = 1
            else:
                header_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(sheet_name)
            # The sheet is new, so headers will go to the first row
            header_row = 1

    # If the header row is 1, it means we need to add headers
    if header_row == 1:
        headers = ["Original Image", "Preprocessed Image", "Converted Text"]
        ws.append(headers)

    # Now append the OCR results starting from the row after headers
    for ocr_result in ocred_tuples:
        org_img, prep_img, text = ocr_result
        # Append the text in the next available row
        ws.append(["", "", text])

        current_row = ws.max_row

        # Embed the original and preprocessed images in their respective cells
        add_image_to_excel_from_memory(
            image=org_img, sheet=ws, row=current_row, col="A"
        )
        add_image_to_excel_from_memory(
            image=prep_img, sheet=ws, row=current_row, col="B"
        )

    # Use the function to adjust the widths
    adjust_column_widths(ws)

    # Set the height of a specific row
    row_height = (
        20 * ENLARGE_FACTOR
    )  # Example height in points; adjust to your needs
    ws.row_dimensions[1].height = row_height

    # If you want to set the height of all rows to the same value:
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = row_height

    # Save the workbook
    wb.save(filename=file_path)


if __name__ == "__main__":

    # Initialize an empty list to collect OCR results
    all_ocr_results = []
    for i in range(1, 2):
        image_path = (
            "/home/iyid/workspaces/price_change_history/"
            f"images/runes/20240309_runes_{i}.png"
        )
        # TODO : Preprocessing enhancement proposition
        # 1. Make the contour boxes thicker/squarer.
        original_img, thresh = preprocess_image_for_contours(
            img_path=image_path
        )
        lines = detect_text_elements(img=original_img, thresh=thresh)

        selected_phrases = []

        for line in lines:
            # line.draw(img=original_img, color=BLUE, display=True)
            for i, phrase in enumerate(line.phrases):
                if i != 1:  # Skip this phrase (a column in the image)
                    # phrase.draw(img=original_img, color=RED, display=True)
                    selected_phrases.append(phrase)
                    for word in phrase.words:
                        orig_word_img, prepr_word_img, word_txt = (
                            word.ocr_text_with_pytesseract(
                                original_img=original_img
                            ).values()
                        )

                        # Collect each set of results in a list
                        all_ocr_results.append(
                            (orig_word_img, prepr_word_img, word_txt)
                        )

    # Now call add_results_to_excel once with all collected OCR results
    add_results_to_excel(
        file_path="runes.xlsx",
        sheet_name="20240302",
        ocred_tuples=all_ocr_results,  # Pass the list of all OCR results
        delete_file_if_exists=True,
    )
